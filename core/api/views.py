import csv
from io import BytesIO, StringIO
import zipfile

from django.conf import settings
from django.core.cache import cache
from django.http import HttpResponse
from openpyxl import load_workbook
from rest_framework.views import APIView

from core.models import GeoPosition, WeatherRecord
from core.tasks import build_weather_statistics_report


class UploadAPI(APIView):

    @staticmethod
    def post(request, *args, **kwargs):
        # TODO: perform some input data validation and handle possible exceptions

        workbook = load_workbook(filename=BytesIO(request.body))

        # Read the first worksheet and discard the other ones
        lines_generator = workbook[workbook.sheetnames[0]].values

        header = next(lines_generator)
        assert header == ('date', 'time', 'geoposition coordinates',
                          'hight under the ground', 'temperature',
                          'wind speed', 'wind vector direction')

        for line in lines_generator:
            record = dict(zip(header, line))

            # Expected format: x,y
            latitude, longitude = tuple(
                map(int, record['geoposition coordinates'].split(','))
            )
            elevation = int(record['hight under the ground'])

            geo_position, _ = GeoPosition.objects.get_or_create(
                latitude=latitude,
                longitude=longitude,
                elevation=elevation,
            )

            WeatherRecord.objects.create(
                geo_position=geo_position,
                date=record['date'],
                time=record['time'],
                temperature=record['temperature'],
                wind_speed=record['wind speed'],
                wind_direction=record['wind vector direction'],
            )

        return HttpResponse('OK', content_type='application/json')


class DownloadAPI(APIView):

    @staticmethod
    def get(request, *args, **kwargs):
        payload = cache.get(settings.UPDATE_WEATHER_STATISTICS_CACHE_KEY)
        if payload is None:
            payload = build_weather_statistics_report()

        zip_file = BytesIO()

        with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zip_out:
            csv_file = StringIO()
            csv.writer(csv_file, dialect='excel').writerows(payload)

            zip_out.writestr('report.csv', csv_file.getvalue())

        response = HttpResponse(zip_file.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=weather_statistics.zip'
        return response
